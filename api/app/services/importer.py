"""Dreamborn collection import: CSV/xlsx parsing + merge/replace UPSERT.

Two Dreamborn export schemas are supported, detected by header:
  A) variant rows (the app's deck/collection scan export):
     Set Number, Card Number, Variant, Count, Name, Color, Rarity
     — one row per (card, variant), Variant is 'normal' or 'foil'
  B) legacy/two-column counts:
     Name, Normal, Foil, [Color, Rarity,] Set, Card Number
"""
import csv
import hashlib
import io

from psycopg.types.json import Jsonb

from . import snapshots
from .matching import Matcher

DIFF_CARD_CAP = 500   # stored per-card diff rows; totals always cover everything

VARIANT_COLUMNS = {"Set Number", "Card Number", "Variant", "Count", "Name"}
LEGACY_COLUMNS = {"Name", "Normal", "Foil", "Set", "Card Number"}


class ImportError_(Exception):
    pass


class DuplicateFileError(Exception):
    def __init__(self, previous_import_id: int):
        self.previous_import_id = previous_import_id
        super().__init__(f"file already imported (import #{previous_import_id})")


def _raw_from_csv(data: bytes) -> tuple[list[str], list[dict]]:
    text = data.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    return list(reader.fieldnames or []), list(reader)


def _raw_from_xlsx(data: bytes) -> tuple[list[str], list[dict]]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    try:
        header = [str(h).strip() if h is not None else "" for h in next(it)]
    except StopIteration:
        raise ImportError_("empty spreadsheet") from None
    rows = []
    for values in it:
        if values is None or all(v is None for v in values):
            continue
        rows.append({header[i]: ("" if v is None else v) for i, v in enumerate(values) if i < len(header)})
    return header, rows


def _to_qty(value) -> int:
    s = str(value).strip()
    if not s:
        return 0
    return max(0, int(float(s)))


def parse_file(filename: str, data: bytes) -> list[dict]:
    """Parse either schema into normalized rows:
    {line, name, set, number, qty_normal, qty_foil, error?}"""
    if filename.lower().endswith((".xlsx", ".xlsm")):
        header, raw = _raw_from_xlsx(data)
    else:
        header, raw = _raw_from_csv(data)
    cols = set(header)

    rows: list[dict] = []
    if VARIANT_COLUMNS.issubset(cols):
        for i, r in enumerate(raw, start=2):
            row = {
                "line": i,
                "name": str(r.get("Name", "")).strip(),
                "set": str(r.get("Set Number", "")).strip(),
                "number": str(r.get("Card Number", "")).strip(),
                "qty_normal": 0,
                "qty_foil": 0,
            }
            variant = str(r.get("Variant", "")).strip().lower()
            try:
                count = _to_qty(r.get("Count", 0))
            except ValueError:
                row["error"] = "bad count"
                rows.append(row)
                continue
            if variant == "normal":
                row["qty_normal"] = count
            elif variant == "foil":
                row["qty_foil"] = count
            else:
                row["error"] = f"unknown variant {variant!r}"
            rows.append(row)
    elif LEGACY_COLUMNS.issubset(cols):
        for i, r in enumerate(raw, start=2):
            row = {
                "line": i,
                "name": str(r.get("Name", "")).strip(),
                "set": str(r.get("Set", "")).strip(),
                "number": str(r.get("Card Number", "")).strip(),
                "qty_normal": 0,
                "qty_foil": 0,
            }
            try:
                row["qty_normal"] = _to_qty(r.get("Normal", 0))
                row["qty_foil"] = _to_qty(r.get("Foil", 0))
            except ValueError:
                row["error"] = "bad quantity"
            rows.append(row)
    else:
        raise ImportError_(
            "unrecognized header: expected Dreamborn variant columns "
            f"{sorted(VARIANT_COLUMNS)} or legacy columns {sorted(LEGACY_COLUMNS)}; got {header}"
        )
    return rows


def _compute_diff(cur, mode: str, counts: dict[str, list[int]]) -> dict:
    """Per-card before→after for this import (works for dry runs too).
    Replace: after = file counts (missing → 0). Merge: after = before + file."""
    cur.execute(
        """SELECT card_id, qty_normal, qty_foil FROM collection
           WHERE qty_normal > 0 OR qty_foil > 0"""
    )
    before = {r["card_id"]: (r["qty_normal"], r["qty_foil"]) for r in cur.fetchall()}
    ids = set(counts) | (set(before) if mode == "replace" else set())

    changed = []
    for card_id in ids:
        bn, bf = before.get(card_id, (0, 0))
        fn, ff = counts.get(card_id, (0, 0))
        an, af = (fn, ff) if mode == "replace" else (bn + fn, bf + ff)
        if (an, af) != (bn, bf):
            changed.append({"card_id": card_id, "before_normal": bn, "before_foil": bf,
                            "after_normal": an, "after_foil": af,
                            "delta": (an + af) - (bn + bf)})

    summary = {
        "cards_changed": len(changed),
        "new_cards": sum(1 for c in changed if c["before_normal"] + c["before_foil"] == 0),
        "removed_cards": sum(1 for c in changed if c["after_normal"] + c["after_foil"] == 0),
        "copies_added": sum(c["delta"] for c in changed if c["delta"] > 0),
        "copies_removed": -sum(c["delta"] for c in changed if c["delta"] < 0),
    }

    changed.sort(key=lambda c: (-abs(c["delta"]), c["card_id"]))
    truncated = max(0, len(changed) - DIFF_CARD_CAP)
    changed = changed[:DIFF_CARD_CAP]
    if changed:
        cur.execute(
            """SELECT c.id, c.full_name, s.code AS set_code, c.collector_number
               FROM cards c JOIN sets s ON s.id = c.set_id WHERE c.id = ANY(%s)""",
            ([c["card_id"] for c in changed],),
        )
        names = {r["id"]: r for r in cur.fetchall()}
        for c in changed:
            n = names.get(c["card_id"], {})
            c["full_name"] = n.get("full_name", "?")
            c["set_code"] = n.get("set_code", "?")
            c["collector_number"] = n.get("collector_number", "?")
    return {"summary": summary, "cards": changed, "truncated": truncated}


def run_import(conn, filename: str, data: bytes, mode: str, dry_run: bool, force: bool,
               note: str = "") -> dict:
    """Match all rows, then apply counts in one transaction. Returns the import report."""
    if mode not in ("merge", "replace"):
        raise ImportError_(f"invalid mode {mode!r}")
    sha = hashlib.sha256(data).hexdigest()
    rows = parse_file(filename, data)

    with conn.cursor() as cur:
        if mode == "merge" and not dry_run and not force:
            cur.execute(
                "SELECT id FROM imports WHERE file_sha256=%s AND NOT dry_run ORDER BY id DESC LIMIT 1",
                (sha,),
            )
            prev = cur.fetchone()
            if prev:
                raise DuplicateFileError(prev["id"])

        matcher = Matcher(cur)
        counts: dict[str, list[int]] = {}   # card_id -> [normal, foil]
        unmatched = []
        for row in rows:
            report = {"row": row["line"], "name": row["name"], "set": row["set"],
                      "number": row["number"]}
            if "error" in row:
                unmatched.append({**report, "reason": row["error"]})
                continue
            card_id, reason = matcher.match(row["set"], row["number"], row["name"])
            if not card_id:
                unmatched.append({**report, "reason": reason})
                continue
            c = counts.setdefault(card_id, [0, 0])
            c[0] += row["qty_normal"]
            c[1] += row["qty_foil"]

        cur.execute("SELECT COALESCE(SUM(qty_normal+qty_foil),0) AS q FROM collection")
        qty_before = cur.fetchone()["q"]

        diff = _compute_diff(cur, mode, counts)   # must read before-state pre-apply

        # Replace-mode safety net: owned cards this file would remove or reduce.
        replace_losses = []
        if mode == "replace":
            cur.execute(
                """SELECT col.card_id, c.full_name, s.code AS set_code,
                          c.collector_number, col.qty_normal, col.qty_foil
                   FROM collection col
                   JOIN cards c ON c.id = col.card_id
                   JOIN sets s ON s.id = c.set_id
                   WHERE col.qty_normal > 0 OR col.qty_foil > 0
                   ORDER BY c.full_name"""
            )
            for r in cur.fetchall():
                fn, ff = counts.get(r["card_id"], (0, 0))
                if fn < r["qty_normal"] or ff < r["qty_foil"]:
                    replace_losses.append({
                        "card_id": r["card_id"], "full_name": r["full_name"],
                        "set_code": r["set_code"], "collector_number": r["collector_number"],
                        "have_normal": r["qty_normal"], "have_foil": r["qty_foil"],
                        "file_normal": fn, "file_foil": ff,
                    })

        added = updated = zeroed = 0
        if not dry_run:
            if mode == "replace":
                cur.execute(
                    "UPDATE collection SET qty_normal=0, qty_foil=0, updated_at=now() "
                    "WHERE qty_normal>0 OR qty_foil>0"
                )
                zeroed = cur.rowcount
            for card_id, (qn, qf) in counts.items():
                if mode == "replace":
                    cur.execute(
                        """INSERT INTO collection (card_id, qty_normal, qty_foil)
                           VALUES (%s, %s, %s)
                           ON CONFLICT (card_id) DO UPDATE
                           SET qty_normal=EXCLUDED.qty_normal, qty_foil=EXCLUDED.qty_foil,
                               updated_at=now()
                           RETURNING (xmax = 0) AS inserted""",
                        (card_id, qn, qf),
                    )
                else:
                    cur.execute(
                        """INSERT INTO collection (card_id, qty_normal, qty_foil)
                           VALUES (%s, %s, %s)
                           ON CONFLICT (card_id) DO UPDATE
                           SET qty_normal=collection.qty_normal+EXCLUDED.qty_normal,
                               qty_foil=collection.qty_foil+EXCLUDED.qty_foil,
                               updated_at=now()
                           RETURNING (xmax = 0) AS inserted""",
                        (card_id, qn, qf),
                    )
                if cur.fetchone()["inserted"]:
                    added += 1
                else:
                    updated += 1

        cur.execute("SELECT COALESCE(SUM(qty_normal+qty_foil),0) AS q FROM collection")
        qty_after = cur.fetchone()["q"] if not dry_run else qty_before

        summary = {
            "added": added, "updated": updated, "zeroed": zeroed,
            "qty_before": int(qty_before), "qty_after": int(qty_after),
            "qty_in_file": sum(qn + qf for qn, qf in counts.values()),
            "lost_cards": len(replace_losses),
        }
        cur.execute(
            """INSERT INTO imports (filename, file_sha256, mode, dry_run, row_count,
                                    matched_rows, unmatched_count, unmatched_rows,
                                    summary, note, diff)
               VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id""",
            (filename, sha, mode, dry_run, len(rows), len(rows) - len(unmatched),
             len(unmatched), Jsonb(unmatched), Jsonb(summary),
             note.strip() or None, Jsonb(diff)),
        )
        import_id = cur.fetchone()["id"]
        if not dry_run:
            # Pin a snapshot to the moment the collection changed, so charts
            # show upload jumps at their true timestamp (not the next daily).
            snapshots.capture(cur, "import", import_id)
    conn.commit()

    return {
        "import_id": import_id, "filename": filename, "mode": mode, "dry_run": dry_run,
        "rows": len(rows), "matched": len(rows) - len(unmatched),
        "unique_cards": len(counts), "unmatched": unmatched,
        "replace_losses": replace_losses, "summary": summary, "diff": diff,
    }
